import boto3
import pandas as pd
from datetime import datetime

VENDOR_ID  = "V07092485"   # change as needed
OUTPUT_FILE = f"vendor_answers_{VENDOR_ID}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

session  = boto3.Session(profile_name="rfp-infra")
dynamodb = session.resource("dynamodb", region_name="us-east-1")
table    = dynamodb.Table("vendor_answers")

# Query all answers for the vendor
resp  = table.query(
    KeyConditionExpression=boto3.dynamodb.conditions.Key("vendor_id").eq(VENDOR_ID)
)
items = resp["Items"]

# Handle DynamoDB pagination
while "LastEvaluatedKey" in resp:
    resp  = table.query(
        KeyConditionExpression=boto3.dynamodb.conditions.Key("vendor_id").eq(VENDOR_ID),
        ExclusiveStartKey=resp["LastEvaluatedKey"]
    )
    items.extend(resp["Items"])

print(f"Total records fetched: {len(items)}")

# Flatten for Excel
rows = []
for item in items:
    sort_key_parts = item.get("sort_key", "#").split("#", 1)
    rows.append({
        "vendor_id":       item.get("vendor_id", ""),
        "questionaire_id": sort_key_parts[0],
        "question_id":     sort_key_parts[1] if len(sort_key_parts) > 1 else "",
        "question_text":   item.get("question_text", ""),
        "answer_value":    item.get("answer_value", ""),
        "answer_state":    item.get("answer_state", ""),
        "confidence":      float(item.get("confidence", 0.0)),
        "citations":       ", ".join(item.get("citations", [])),
        "rationale":       item.get("rationale", ""),
        "suggestion":      item.get("suggestion", ""),
        "updated_at":      item.get("updated_at", ""),
    })

# Sort by questionaire_id then question_id
rows.sort(key=lambda x: (x["questionaire_id"], x["question_id"]))

df = pd.DataFrame(rows)

# Write to Excel with formatting
with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:
    df.to_excel(writer, index=False, sheet_name="Vendor Answers")

    ws = writer.sheets["Vendor Answers"]

    # Auto-fit column widths
    for col in ws.columns:
        max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)

    # Bold header row
    from openpyxl.styles import Font, PatternFill, Alignment
    header_fill = PatternFill(start_color="1C4548", end_color="1C4548", fill_type="solid")
    for cell in ws[1]:
        cell.font      = Font(bold=True, color="FFFFFF")
        cell.fill      = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Colour-code answer_state column (column F = index 6)
    state_colours = {
        "SUPPORTED":             "C6EFCE",   # green
        "INFERRED":              "FFEB9C",   # amber
        "INSUFFICIENT_EVIDENCE": "FFC7CE",   # red
    }
    state_col_idx = df.columns.get_loc("answer_state") + 1
    for row in ws.iter_rows(min_row=2, min_col=state_col_idx, max_col=state_col_idx):
        for cell in row:
            colour = state_colours.get(str(cell.value), "FFFFFF")
            cell.fill = PatternFill(start_color=colour, end_color=colour, fill_type="solid")

print(f"Exported to: {OUTPUT_FILE}")
